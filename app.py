import os
import re
import json
from datetime import datetime
from typing import Optional, Dict, Tuple, List, Any

from flask import Flask, request, jsonify
from flask_cors import CORS

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

# =========================================
# CONFIG
# =========================================
HEADERS = [
    "Card_Last_4",
    "Bank",
    "Card_Label",
    "Date",
    "Currency",
    "Amount",
    "FX_Rate_To_SGD",
    "Amount_SGD",
    "FX_Rate_Date",
    "FX_Source",
    "Description",
    "Raw_SMS",
    "Created_At",
]

BASE_CURRENCY = "SGD"
FX_SOURCE_NAME = "Frankfurter"
FX_API_BASE_URL = "https://api.frankfurter.dev/v1"


def resolve_data_dir() -> str:
    """
    Resolve a writable data directory.

    Order:
    1. DATA_DIR environment variable if explicitly provided
    2. /app/data if it already exists (Render persistent disk mount)
    3. ./data under current working directory as safe fallback
    """
    env_dir = os.getenv("DATA_DIR")
    if env_dir:
        return env_dir

    render_disk = "/app/data"
    if os.path.exists(render_disk):
        return render_disk

    return os.path.join(os.getcwd(), "data")


DATA_DIR = resolve_data_dir()
EXCEL_FILE = os.path.join(DATA_DIR, "transactions.xlsx")
STATE_FILE = os.path.join(DATA_DIR, "app_state.json")
FX_CACHE_FILE = os.path.join(DATA_DIR, "fx_cache.json")


def refresh_paths() -> None:
    global EXCEL_FILE, STATE_FILE, FX_CACHE_FILE
    EXCEL_FILE = os.path.join(DATA_DIR, "transactions.xlsx")
    STATE_FILE = os.path.join(DATA_DIR, "app_state.json")
    FX_CACHE_FILE = os.path.join(DATA_DIR, "fx_cache.json")


# =========================================
# FILE SETUP
# =========================================
def ensure_data_folder() -> None:
    """
    Create the data folder safely.
    If the configured folder is not writable, fall back to ./data.
    """
    global DATA_DIR
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
    except PermissionError:
        DATA_DIR = os.path.join(os.getcwd(), "data")
        os.makedirs(DATA_DIR, exist_ok=True)
        refresh_paths()


def create_excel_if_missing() -> None:
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)


def create_state_if_missing() -> None:
    if not os.path.exists(STATE_FILE):
        state = {"last_reset_month": ""}
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2)


def create_fx_cache_if_missing() -> None:
    if not os.path.exists(FX_CACHE_FILE):
        with open(FX_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f, indent=2)


def get_excel_headers() -> List[str]:
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Transactions"]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if not first_row:
        return []

    return [str(col) if col is not None else "" for col in first_row]


def rebuild_excel_with_new_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)
    print("[INFO] Excel file rebuilt with latest headers.")


def initialize_files() -> None:
    ensure_data_folder()
    create_excel_if_missing()
    create_state_if_missing()
    create_fx_cache_if_missing()

    existing_headers = get_excel_headers()
    if existing_headers and existing_headers != HEADERS:
        print("[WARNING] Existing Excel format is outdated or different.")
        print("[WARNING] Rebuilding transactions.xlsx with latest headers.")
        rebuild_excel_with_new_headers()


# =========================================
# STATE / RESET
# =========================================
def read_state() -> Dict[str, Any]:
    create_state_if_missing()
    with open(STATE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def write_state(state: Dict[str, Any]) -> None:
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def read_fx_cache() -> Dict[str, Any]:
    create_fx_cache_if_missing()
    with open(FX_CACHE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def write_fx_cache(cache: Dict[str, Any]) -> None:
    with open(FX_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)


def clear_excel_transactions() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)


def ensure_monthly_reset(today: Optional[datetime] = None) -> None:
    if today is None:
        today = datetime.now()

    current_month_key = today.strftime("%Y-%m")
    state = read_state()
    last_reset_month = state.get("last_reset_month", "")

    if today.day == 1 and last_reset_month != current_month_key:
        clear_excel_transactions()
        state["last_reset_month"] = current_month_key
        write_state(state)
        print(f"[INFO] Auto-reset completed for month: {current_month_key}")


# =========================================
# PARSING
# =========================================
def parse_date_to_datetime(date_text: str) -> Optional[datetime]:
    formats = ["%d/%m/%y", "%d/%m/%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_text.strip(), fmt)
        except ValueError:
            continue
    return None


def extract_date(text: str) -> Optional[str]:
    match = re.search(r"\b(\d{2}/\d{2}/(?:\d{2}|\d{4}))\b", text)
    return match.group(1) if match else None


def extract_amount_and_currency(text: str) -> Tuple[Optional[str], Optional[float]]:
    match = re.search(r"\b([A-Z]{3})\s*([\d,]+(?:\.\d{1,2})?)\b", text, re.IGNORECASE)
    if not match:
        return None, None

    currency = match.group(1).upper()
    amount_str = match.group(2).replace(",", "")

    try:
        return currency, float(amount_str)
    except ValueError:
        return currency, None


def extract_description(text: str) -> Optional[str]:
    match = re.search(
        r"\bat\s+(.+?)(?:\.\s|\s+[A-Z]{3}\b|\s+card\s+ending\s+with\b|\s+ending\s+with\b|$)",
        text,
        re.IGNORECASE,
    )
    return match.group(1).strip(" .") if match else None


def extract_card_last_4(text: str) -> Optional[str]:
    match = re.search(r"ending(?:\s+with)?\s+(\d{4})\b", text, re.IGNORECASE)
    return match.group(1) if match else None


def parse_sms_content(text: str) -> Dict[str, Optional[Any]]:
    date_text = extract_date(text)
    currency, amount = extract_amount_and_currency(text)
    description = extract_description(text)
    card_last_4 = extract_card_last_4(text)

    return {
        "date": date_text,
        "currency": currency,
        "amount": amount,
        "description": description,
        "card_last_4": card_last_4,
    }


# =========================================
# FX CONVERSION
# =========================================
def to_api_date(date_text: str) -> Optional[str]:
    parsed = parse_date_to_datetime(date_text)
    return parsed.strftime("%Y-%m-%d") if parsed else None


def build_fx_cache_key(date_yyyy_mm_dd: str, from_currency: str, to_currency: str) -> str:
    return f"{date_yyyy_mm_dd}|{from_currency.upper()}|{to_currency.upper()}"


def get_historical_fx_rate_to_sgd(from_currency: str, transaction_date: str) -> Dict[str, Any]:
    from_currency = from_currency.upper()

    if from_currency == BASE_CURRENCY:
        api_date = to_api_date(transaction_date)
        return {
            "success": True,
            "rate": 1.0,
            "fx_rate_date": api_date,
            "source": FX_SOURCE_NAME,
            "message": "Base currency is already SGD.",
        }

    api_date = to_api_date(transaction_date)
    if not api_date:
        return {
            "success": False,
            "rate": None,
            "fx_rate_date": None,
            "source": FX_SOURCE_NAME,
            "message": "Invalid transaction date for FX conversion.",
        }

    cache = read_fx_cache()
    cache_key = build_fx_cache_key(api_date, from_currency, BASE_CURRENCY)

    if cache_key in cache:
        cached = cache[cache_key]
        return {
            "success": True,
            "rate": cached.get("rate"),
            "fx_rate_date": cached.get("fx_rate_date"),
            "source": cached.get("source", FX_SOURCE_NAME),
            "message": "Loaded FX rate from cache.",
        }

    url = f"{FX_API_BASE_URL}/{api_date}"
    params = {"base": from_currency, "symbols": BASE_CURRENCY}

    try:
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()

        rates = data.get("rates", {})
        rate = rates.get(BASE_CURRENCY)
        fx_rate_date = data.get("date")

        if rate is None:
            return {
                "success": False,
                "rate": None,
                "fx_rate_date": fx_rate_date,
                "source": FX_SOURCE_NAME,
                "message": f"No FX rate returned for {from_currency}->{BASE_CURRENCY}.",
            }

        cache[cache_key] = {
            "rate": float(rate),
            "fx_rate_date": fx_rate_date,
            "source": FX_SOURCE_NAME,
        }
        write_fx_cache(cache)

        return {
            "success": True,
            "rate": float(rate),
            "fx_rate_date": fx_rate_date,
            "source": FX_SOURCE_NAME,
            "message": "FX rate fetched successfully.",
        }

    except requests.RequestException as e:
        return {
            "success": False,
            "rate": None,
            "fx_rate_date": None,
            "source": FX_SOURCE_NAME,
            "message": f"FX API request failed: {e}",
        }
    except ValueError as e:
        return {
            "success": False,
            "rate": None,
            "fx_rate_date": None,
            "source": FX_SOURCE_NAME,
            "message": f"FX API response parse failed: {e}",
        }


def convert_amount_to_sgd(amount: float, currency: str, transaction_date: str) -> Dict[str, Any]:
    fx_result = get_historical_fx_rate_to_sgd(currency, transaction_date)

    if not fx_result["success"]:
        return {
            "success": False,
            "amount_sgd": None,
            "fx_rate": None,
            "fx_rate_date": fx_result.get("fx_rate_date"),
            "fx_source": fx_result.get("source"),
            "message": fx_result.get("message"),
        }

    fx_rate = fx_result["rate"]
    amount_sgd = round(float(amount) * float(fx_rate), 2)

    return {
        "success": True,
        "amount_sgd": amount_sgd,
        "fx_rate": fx_rate,
        "fx_rate_date": fx_result.get("fx_rate_date"),
        "fx_source": fx_result.get("source"),
        "message": fx_result.get("message"),
    }


# =========================================
# EXCEL OPERATIONS
# =========================================
def append_transaction(row_data: List[Any]) -> None:
    create_excel_if_missing()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Transactions"]
    ws.append(row_data)
    wb.save(EXCEL_FILE)


def load_transactions() -> List[Dict[str, Any]]:
    create_excel_if_missing()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Transactions"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) == 1:
        return []

    headers = rows[0]
    data_rows = rows[1:]

    transactions = []
    for row in data_rows:
        if any(cell is not None for cell in row):
            transactions.append(dict(zip(headers, row)))

    return transactions


def load_transactions_df() -> pd.DataFrame:
    transactions = load_transactions()
    if not transactions:
        return pd.DataFrame(columns=HEADERS)

    df = pd.DataFrame(transactions)

    for col in HEADERS:
        if col not in df.columns:
            df[col] = None

    df = df[HEADERS]

    numeric_cols = ["Amount", "FX_Rate_To_SGD", "Amount_SGD"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


# =========================================
# TOTALS / FILTERING
# =========================================
def get_current_month_total(transactions: List[Dict[str, Any]]) -> Dict[str, float]:
    totals: Dict[str, float] = {}
    now = datetime.now()

    for txn in transactions:
        date_str = txn.get("Date")
        currency = txn.get("Currency")
        amount = txn.get("Amount")

        if not date_str or amount is None or not currency:
            continue

        parsed_date = parse_date_to_datetime(str(date_str))
        if parsed_date is None:
            continue

        if parsed_date.year == now.year and parsed_date.month == now.month:
            totals[currency] = totals.get(currency, 0.0) + float(amount)

    return totals


def get_current_month_total_sgd(transactions: List[Dict[str, Any]]) -> float:
    total_sgd = 0.0
    now = datetime.now()

    for txn in transactions:
        date_str = txn.get("Date")
        amount_sgd = txn.get("Amount_SGD")

        if not date_str or amount_sgd is None:
            continue

        parsed_date = parse_date_to_datetime(str(date_str))
        if parsed_date is None:
            continue

        if parsed_date.year == now.year and parsed_date.month == now.month:
            total_sgd += float(amount_sgd)

    return round(total_sgd, 2)


def get_transactions_for_card(card_last_4: str) -> List[Dict[str, Any]]:
    df = load_transactions_df()
    if df.empty:
        return []
    card_df = df[df["Card_Last_4"].astype(str) == str(card_last_4)]
    return card_df.to_dict("records")


def get_current_month_transactions_for_card(card_last_4: str) -> List[Dict[str, Any]]:
    df = load_transactions_df()
    if df.empty:
        return []

    card_df = df[df["Card_Last_4"].astype(str) == str(card_last_4)]
    if card_df.empty:
        return []

    now = datetime.now()

    def is_current_month(date_str: str) -> bool:
        parsed = parse_date_to_datetime(str(date_str))
        return bool(parsed and parsed.year == now.year and parsed.month == now.month)

    current_month_df = card_df[card_df["Date"].apply(is_current_month)]
    return current_month_df.to_dict("records")


def get_monthly_totals_by_card() -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    now = datetime.now()

    df = load_transactions_df()
    if df.empty:
        return result

    all_cards = sorted(df["Card_Last_4"].dropna().astype(str).unique().tolist())

    for card_last_4 in all_cards:
        card_df = df[df["Card_Last_4"].astype(str) == card_last_4].copy()

        currency_totals: Dict[str, float] = {}
        amount_sgd_total = 0.0

        for _, row in card_df.iterrows():
            date_str = row.get("Date")
            currency = row.get("Currency")
            amount = row.get("Amount")
            amount_sgd = row.get("Amount_SGD")

            parsed_date = parse_date_to_datetime(str(date_str))
            if parsed_date is None:
                continue

            if parsed_date.year == now.year and parsed_date.month == now.month:
                if pd.notna(amount) and currency:
                    currency_totals[currency] = currency_totals.get(currency, 0.0) + float(amount)
                if pd.notna(amount_sgd):
                    amount_sgd_total += float(amount_sgd)

        result[card_last_4] = {
            "currency_totals": currency_totals,
            "amount_sgd_total": round(amount_sgd_total, 2),
        }

    return result


def get_overall_totals_by_card_all_time() -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}

    df = load_transactions_df()
    if df.empty:
        return result

    all_cards = sorted(df["Card_Last_4"].dropna().astype(str).unique().tolist())

    for card_last_4 in all_cards:
        card_df = df[df["Card_Last_4"].astype(str) == card_last_4].copy()

        currency_totals: Dict[str, float] = {}
        amount_sgd_total = 0.0

        for _, row in card_df.iterrows():
            currency = row.get("Currency")
            amount = row.get("Amount")
            amount_sgd = row.get("Amount_SGD")

            if pd.notna(amount) and currency:
                currency_totals[currency] = currency_totals.get(currency, 0.0) + float(amount)

            if pd.notna(amount_sgd):
                amount_sgd_total += float(amount_sgd)

        result[card_last_4] = {
            "currency_totals": currency_totals,
            "amount_sgd_total": round(amount_sgd_total, 2),
            "card_label": card_df.iloc[0]["Card_Label"] if not card_df.empty else f"Card - {card_last_4}",
            "bank": card_df.iloc[0]["Bank"] if not card_df.empty else "Unknown",
        }

    return result


# =========================================
# BUSINESS LOGIC
# =========================================
def detect_bank_from_sms(sms_content: str, card_last_4: str) -> str:
    sms_upper = sms_content.upper()

    if "UOB" in sms_upper:
        return "UOB"
    if "OCBC" in sms_upper:
        return "OCBC"
    if "DBS" in sms_upper or "POSB" in sms_upper:
        return "DBS"
    if "CITI" in sms_upper or "CITIBANK" in sms_upper:
        return "CITIBANK"
    if "MAYBANK" in sms_upper:
        return "MAYBANK"
    if "SCB" in sms_upper or "STANDARD CHARTERED" in sms_upper:
        return "Standard Chartered"
    if "HSBC" in sms_upper:
        return "HSBC"

    return "Unknown"


def get_card_info(card_last_4: str, sms_content: str) -> Dict[str, str]:
    bank = detect_bank_from_sms(sms_content, card_last_4)
    label = f"{bank} - {card_last_4}" if bank != "Unknown" else f"Card - {card_last_4}"
    return {"bank": bank, "label": label}


def submit_transaction(sms_content: str) -> Dict[str, Any]:
    initialize_files()
    ensure_monthly_reset()

    sms_content = sms_content.strip()

    if not sms_content:
        return {"success": False, "message": "SMS content is empty."}

    parsed = parse_sms_content(sms_content)

    if not parsed["date"]:
        return {"success": False, "message": "Could not detect date."}

    if not parsed["currency"] or parsed["amount"] is None:
        return {"success": False, "message": "Could not detect amount/currency."}

    if not parsed["description"]:
        return {"success": False, "message": "Could not detect description after 'at'."}

    if not parsed["card_last_4"]:
        return {"success": False, "message": "Could not detect 4-digit card number after 'ending with'."}

    card_info = get_card_info(parsed["card_last_4"], sms_content)

    conversion = convert_amount_to_sgd(
        amount=float(parsed["amount"]),
        currency=str(parsed["currency"]),
        transaction_date=str(parsed["date"]),
    )

    if not conversion["success"]:
        return {
            "success": False,
            "message": f"Transaction parsed but FX conversion failed. {conversion['message']}",
            "parsed": parsed,
            "card_info": card_info,
        }

    row = [
        parsed["card_last_4"],
        card_info["bank"],
        card_info["label"],
        parsed["date"],
        parsed["currency"],
        parsed["amount"],
        conversion["fx_rate"],
        conversion["amount_sgd"],
        conversion["fx_rate_date"],
        conversion["fx_source"],
        parsed["description"],
        sms_content,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    append_transaction(row)

    return {
        "success": True,
        "message": "Transaction saved successfully.",
        "parsed": parsed,
        "card_info": card_info,
        "conversion": conversion,
    }


# =========================================
# ROUTES
# =========================================
@app.route("/", methods=["GET", "HEAD"])
def home():
    return jsonify(
        {
            "success": True,
            "service": "credit-card-sms-tracker",
            "status": "running",
            "data_dir": DATA_DIR,
            "timestamp": datetime.now().isoformat(),
            "endpoints": [
                "/health",
                "/api/transactions",
                "/api/transactions/<card_last_4>",
                "/api/transactions/current-month/<card_last_4>",
                "/api/submit",
                "/api/totals/monthly",
                "/api/totals/monthly/by-card",
                "/api/totals/all-time",
                "/api/cards",
                "/api/stats",
                "/api/parse",
                "/api/reset",
            ],
        }
    ), 200


@app.route("/favicon.ico", methods=["GET"])
def favicon():
    return "", 204


@app.route("/health", methods=["GET"])
def health_check():
    return jsonify(
        {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "data_dir": DATA_DIR,
        }
    ), 200


@app.route("/api/transactions", methods=["GET"])
def get_all_transactions():
    try:
        transactions = load_transactions()
        return jsonify({"success": True, "count": len(transactions), "transactions": transactions}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/transactions/<card_last_4>", methods=["GET"])
def get_transactions_by_card(card_last_4: str):
    try:
        transactions = get_transactions_for_card(card_last_4)
        return jsonify(
            {
                "success": True,
                "card_last_4": card_last_4,
                "count": len(transactions),
                "transactions": transactions,
            }
        ), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/transactions/current-month/<card_last_4>", methods=["GET"])
def get_current_month_transactions_by_card(card_last_4: str):
    try:
        transactions = get_current_month_transactions_for_card(card_last_4)
        return jsonify(
            {
                "success": True,
                "card_last_4": card_last_4,
                "count": len(transactions),
                "transactions": transactions,
            }
        ), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/totals/monthly", methods=["GET"])
def get_monthly_totals():
    try:
        transactions = load_transactions()
        currency_totals = get_current_month_total(transactions)
        sgd_total = get_current_month_total_sgd(transactions)

        return jsonify(
            {
                "success": True,
                "currency_totals": currency_totals,
                "sgd_total": sgd_total,
                "month": datetime.now().strftime("%Y-%m"),
            }
        ), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/totals/monthly/by-card", methods=["GET"])
def get_monthly_totals_by_card_endpoint():
    try:
        totals = get_monthly_totals_by_card()
        df = load_transactions_df()

        result = {}
        for card_last_4, card_data in totals.items():
            card_df = df[df["Card_Last_4"].astype(str) == str(card_last_4)]
            card_label = card_df.iloc[0]["Card_Label"] if not card_df.empty else f"Card - {card_last_4}"
            bank = card_df.iloc[0]["Bank"] if not card_df.empty else "Unknown"

            result[card_last_4] = {
                "card_label": card_label,
                "bank": bank,
                "currency_totals": card_data["currency_totals"],
                "amount_sgd_total": card_data["amount_sgd_total"],
            }

        return jsonify({"success": True, "totals": result, "month": datetime.now().strftime("%Y-%m")}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/totals/all-time", methods=["GET"])
def get_all_time_totals():
    try:
        totals = get_overall_totals_by_card_all_time()
        return jsonify({"success": True, "totals": totals}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/submit", methods=["POST"])
def submit_transaction_api():
    try:
        data = request.get_json(silent=True) or {}
        sms_content = str(data.get("sms_content", "")).strip()

        if not sms_content:
            return jsonify({"success": False, "message": "SMS content is required"}), 400

        result = submit_transaction(sms_content)
        return (jsonify(result), 200) if result["success"] else (jsonify(result), 400)

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/reset", methods=["POST"])
def reset_transactions_api():
    try:
        data = request.get_json(silent=True) or {}
        confirm = str(data.get("confirm", ""))

        if confirm != "yes":
            return jsonify(
                {
                    "success": False,
                    "message": "Reset not confirmed. Please set confirm to 'yes'",
                }
            ), 400

        clear_excel_transactions()

        state = read_state()
        state["last_reset_month"] = ""
        write_state(state)

        return jsonify({"success": True, "message": "All transactions have been cleared"}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/cards", methods=["GET"])
def get_all_cards():
    try:
        df = load_transactions_df()
        if df.empty:
            return jsonify({"success": True, "cards": []}), 200

        cards = []
        for card_last_4 in df["Card_Last_4"].dropna().astype(str).unique():
            card_df = df[df["Card_Last_4"].astype(str) == card_last_4]
            cards.append(
                {
                    "card_last_4": card_last_4,
                    "card_label": card_df.iloc[0]["Card_Label"],
                    "bank": card_df.iloc[0]["Bank"],
                    "transaction_count": int(len(card_df)),
                }
            )

        return jsonify({"success": True, "cards": sorted(cards, key=lambda x: x["card_last_4"])}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/stats", methods=["GET"])
def get_stats():
    try:
        df = load_transactions_df()

        if df.empty:
            return jsonify(
                {
                    "success": True,
                    "stats": {
                        "total_transactions": 0,
                        "total_amount_sgd": 0,
                        "unique_cards": 0,
                        "current_month_transactions": 0,
                        "current_month_amount_sgd": 0,
                    },
                }
            ), 200

        now = datetime.now()

        def is_current_month(date_value: Any) -> bool:
            parsed = parse_date_to_datetime(str(date_value))
            return bool(parsed and parsed.month == now.month and parsed.year == now.year)

        current_month_df = df[df["Date"].apply(is_current_month)]

        stats = {
            "total_transactions": int(len(df)),
            "total_amount_sgd": round(float(df["Amount_SGD"].fillna(0).sum()), 2),
            "unique_cards": int(df["Card_Last_4"].nunique()),
            "current_month_transactions": int(len(current_month_df)),
            "current_month_amount_sgd": round(float(current_month_df["Amount_SGD"].fillna(0).sum()), 2),
        }

        return jsonify({"success": True, "stats": stats}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/parse", methods=["POST"])
def parse_sms_api():
    try:
        data = request.get_json(silent=True) or {}
        sms_content = str(data.get("sms_content", "")).strip()

        if not sms_content:
            return jsonify({"success": False, "message": "SMS content is required"}), 400

        parsed = parse_sms_content(sms_content)
        bank = detect_bank_from_sms(sms_content, parsed.get("card_last_4") or "")

        return jsonify({"success": True, "parsed": parsed, "detected_bank": bank}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# Initialize files when module is imported by Gunicorn
initialize_files()

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(debug=True, host="0.0.0.0", port=port)
